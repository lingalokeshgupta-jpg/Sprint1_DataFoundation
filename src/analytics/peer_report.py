import sqlite3
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

conn = sqlite3.connect("db/nifty100.db")

financial_ratios = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn
)

peer_groups = pd.read_sql(
    "SELECT * FROM peer_groups",
    conn
)

peer_percentiles = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn
)

companies = pd.read_sql(
    """
    SELECT
        id AS company_id,
        company_name
    FROM companies
    """,
    conn
)

merged = financial_ratios.merge(

    peer_groups,

    on="company_id",

    how="left"

)

merged["peer_group_name"] = (

    merged["peer_group_name"]

    .fillna("No peer group assigned")

)


merged = merged.merge(
    companies,
    on="company_id",
    how="left"
)
writer = pd.ExcelWriter(

    "output/peer_comparison.xlsx",

    engine="openpyxl"

)

peer_list = (

    merged["peer_group_name"]

    .dropna()

    .unique()

)

for peer in peer_list:

    sheet = merged[
        merged["peer_group_name"] == peer
    ].copy()

    percentile_sheet = peer_percentiles[
        peer_percentiles["peer_group_name"] == peer
    ]

    percentile_pivot = (
        percentile_sheet
        .pivot_table(
            index=["company_id", "year"],
            columns="metric",
            values="percentile_rank"
        )
        .reset_index()
    )

    percentile_pivot.columns = [
        str(col) + "_pct"
        if col not in ["company_id", "year"]
        else col
        for col in percentile_pivot.columns
    ]

    sheet = sheet.merge(
        percentile_pivot,
        on=["company_id", "year"],
        how="left"
    )

    sheet.to_excel(
        writer,
        sheet_name=peer[:31],
        index=False
    )

    ws = writer.sheets[peer[:31]]

    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 2
    gold_fill = PatternFill(
        fill_type="solid",
        start_color="FFD966",
        end_color="FFD966"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="92D050",
        end_color="92D050"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFD966",
        end_color="FFD966"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="F8696B",
        end_color="F8696B"
    )

    headers = {}

    for cell in ws[1]:
        headers[cell.value] = cell.column

    benchmark_col = headers["is_benchmark"]

    for row in range(2, ws.max_row + 1):

        benchmark = ws.cell(
            row=row,
            column=benchmark_col
        ).value

        if benchmark == 1:

            for cell in ws[row]:

                cell.fill = gold_fill

    median = sheet.median(numeric_only=True)
    ws.append(["Median"] + median.tolist())
    last_row = ws.max_row

    for cell in ws[last_row]:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )
writer.close()

print("Peer comparison workbook created successfully!")

print(merged.shape)

print(
    merged["peer_group_name"]

    .value_counts()
)

